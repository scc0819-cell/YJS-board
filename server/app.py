#!/usr/bin/env python3
"""
昱金生能源集團 - 員工日報系統 v3.0
改進：
1. 登入驗證（帳號密碼保護）
2. 案場下拉選單（不需手打）
3. AI 自動分析（提交即觸發）
4. 歷史查詢功能
5. 草稿自動儲存
6. 自動提醒（超時未交）
7. 員工管理後台
"""

from flask import Flask, request, render_template, redirect, url_for, flash, jsonify, session, send_file, abort
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import os
import json
import hashlib
from pathlib import Path
import threading
import subprocess
import io
import sqlite3
import re
import secrets
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

app = Flask(__name__)
# NOTE: 正式環境請改用環境變數或外部設定檔提供 secret key
app.secret_key = os.environ.get('YJS_DAILY_REPORT_SECRET_KEY', 'yjsenergy_v3_2026_ultra_secret')
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    # 限制單次請求大小（含表單欄位 + 附件），避免意外上傳過大檔案拖垮服務
    MAX_CONTENT_LENGTH=100 * 1024 * 1024,  # 100MB
)
CORS(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '請先登入'

# ===================== 路徑設定 =====================
BASE_DIR = Path('/home/yjsclaw/.openclaw/workspace')
REPORTS_DIR = BASE_DIR / 'daily_reports'
ATTACHMENTS_DIR = BASE_DIR / 'daily_report_attachments'
SCRIPTS_DIR = BASE_DIR / 'scripts'
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)

# ===================== 帳號資料 =====================
# ✅ 改善重點：
# - 不再把預設密碼顯示在畫面上
# - 支援登入後自行修改密碼
# - 使用者資料改為落地到 JSON（重啟不會消失）
# - 密碼雜湊改用 werkzeug（含 salt），避免自製 SHA256

DATA_DIR = BASE_DIR / 'server' / 'data'
USERS_FILE = DATA_DIR / 'users.json'
DATA_DIR.mkdir(parents=True, exist_ok=True)

# 初始帳號（首次建立 users.json 時才會使用）
# 注意：員工初始密碼不在畫面顯示；首次登入會被要求修改密碼。
SEED_USERS = {
    "admin": {
        "id": "admin",
        "name": "宋董事長",
        "role": "admin",
        "enabled": True,
        "password_hash": generate_password_hash("yjsenergy2026"),
        "password_temp": False,
    },
    "chen_ming_de": {"id": "chen_ming_de", "name": "陳明德", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "yang_zong_wei": {"id": "yang_zong_wei", "name": "楊宗衛", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "li_ya_ting": {"id": "li_ya_ting", "name": "李雅婷", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "hong_shu_rong": {"id": "hong_shu_rong", "name": "洪淑嫆", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "chen_gu_bin": {"id": "chen_gu_bin", "name": "陳谷濱", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "zhang_yi_chuan": {"id": "zhang_yi_chuan", "name": "張億峖", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "lin_kun_yi": {"id": "lin_kun_yi", "name": "林坤誼", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "huang_zhen_hao": {"id": "huang_zhen_hao", "name": "黃振豪", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
    "xu_hui_ling": {"id": "xu_hui_ling", "name": "許惠玲", "role": "employee", "enabled": True, "password_hash": generate_password_hash("1234"), "password_temp": True},
}


def normalize_users(users: dict):
    """補齊 users.json 欄位（向後相容），避免版本升級後缺欄位出錯。"""
    changed = False
    for uid, u in (users or {}).items():
        if not isinstance(u, dict):
            users[uid] = {'id': uid, 'name': uid, 'role': 'employee', 'enabled': True, 'password_temp': True}
            changed = True
            continue
        if 'id' not in u:
            u['id'] = uid; changed = True
        if 'name' not in u:
            u['name'] = uid; changed = True
        if 'role' not in u:
            u['role'] = 'employee'; changed = True
        if 'enabled' not in u:
            u['enabled'] = True; changed = True
        if 'password_temp' not in u:
            u['password_temp'] = False; changed = True
        users[uid] = u
    return users, changed


def load_users():
    if USERS_FILE.exists():
        try:
            users = json.loads(USERS_FILE.read_text(encoding='utf-8'))
            users, changed = normalize_users(users)
            if changed:
                save_users(users)
            return users
        except Exception:
            # 檔案損壞時回退到 seed（避免整個系統起不來）
            users = dict(SEED_USERS)
            users, _ = normalize_users(users)
            return users

    # 首次建立
    users = dict(SEED_USERS)
    users, _ = normalize_users(users)
    USERS_FILE.write_text(json.dumps(users, ensure_ascii=False, indent=2), encoding='utf-8')
    return users


def save_users(users: dict):
    tmp = USERS_FILE.with_suffix('.json.tmp')
    tmp.write_text(json.dumps(users, ensure_ascii=False, indent=2), encoding='utf-8')
    tmp.replace(USERS_FILE)


def public_user(u: dict) -> dict:
    return {
        'id': u.get('id'),
        'name': u.get('name'),
        'role': u.get('role'),
        'enabled': bool(u.get('enabled', True)),
    }


USERS = load_users()


def refresh_employee_cache():
    global EMPLOYEES
    EMPLOYEES = [public_user(v) for v in USERS.values() if v.get('role') == 'employee' and bool(v.get('enabled', True))]
    EMPLOYEES.sort(key=lambda x: x.get('name') or '')


refresh_employee_cache()


# ===================== 安全與稽核 =====================
AUTH_STATE_FILE = DATA_DIR / 'auth_state.json'
AUDIT_LOG_FILE = DATA_DIR / 'audit.jsonl'


def _now_ts():
    return int(datetime.now().timestamp())


def load_auth_state():
    if AUTH_STATE_FILE.exists():
        try:
            return json.loads(AUTH_STATE_FILE.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}


def save_auth_state(state: dict):
    tmp = AUTH_STATE_FILE.with_suffix('.json.tmp')
    tmp.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding='utf-8')
    tmp.replace(AUTH_STATE_FILE)


def client_ip():
    xf = request.headers.get('X-Forwarded-For', '')
    if xf:
        return xf.split(',')[0].strip()
    return request.remote_addr or 'unknown'


def audit(event: str, actor_id: str = '', target_id: str = '', detail: dict | None = None):
    """寫入稽核紀錄（JSONL）。避免寫入密碼等敏感資訊。"""
    rec = {
        'ts': _now_ts(),
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'event': event,
        'actor_id': actor_id,
        'target_id': target_id,
        'ip': client_ip(),
        'path': request.path,
        'detail': detail or {},
    }
    try:
        with open(AUDIT_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        pass


def tail_jsonl(p: Path, max_lines: int = 400):
    if not p.exists():
        return []
    try:
        lines = p.read_text(encoding='utf-8', errors='ignore').splitlines()[-max_lines:]
        out = []
        for ln in lines:
            try:
                out.append(json.loads(ln))
            except Exception:
                continue
        return out
    except Exception:
        return []


# ===================== 資料庫（SQLite） =====================
DB_FILE = DATA_DIR / 'app.db'

# 里程碑（你指定 C：里程碑 + 百分比都要）
MILESTONE_STAGES = [
    '現勘', '投標', '得標', '設計', '送審', '施工', '完工掛表', '維運'
]

RISK_CATEGORIES = ['工期', '成本', '協調', '合約', '品質', '其他']


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with get_db() as db:
        db.execute('''
            CREATE TABLE IF NOT EXISTS reports (
              report_key TEXT PRIMARY KEY,
              report_date TEXT,
              employee_id TEXT,
              submitted_at TEXT,
              report_json TEXT
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS work_items (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_key TEXT,
              case_id TEXT,
              work_type TEXT,
              progress REAL,
              hours REAL,
              work_content TEXT,
              output TEXT,
              status TEXT
            )
        ''')

        db.execute('''
            CREATE TABLE IF NOT EXISTS plan_items (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_key TEXT,
              case_id TEXT,
              plan_content TEXT,
              plan_hours REAL,
              need_support TEXT
            )
        ''')

        db.execute('''
            CREATE TABLE IF NOT EXISTS attachments (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_key TEXT,
              report_date TEXT,
              employee_id TEXT,
              kind TEXT,
              original_name TEXT,
              stored_name TEXT,
              rel_path TEXT,
              size_bytes INTEGER,
              uploaded_at TEXT
            )
        ''')

        db.execute('''
            CREATE TABLE IF NOT EXISTS risk_items (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              report_key TEXT,
              report_date TEXT,
              employee_id TEXT,
              case_id TEXT,
              category TEXT,
              level TEXT,
              risk_desc TEXT,
              risk_impact TEXT,
              need_help TEXT,
              status TEXT DEFAULT 'open',
              owner_id TEXT DEFAULT '',
              due_date TEXT DEFAULT '',
              created_at TEXT
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS cases (
              case_id TEXT PRIMARY KEY,
              name TEXT,
              type TEXT,
              enabled INTEGER DEFAULT 1,
              created_at TEXT,
              updated_at TEXT
            )
        ''')

        db.execute('''
            CREATE TABLE IF NOT EXISTS case_aliases (
              alias TEXT PRIMARY KEY,
              case_id TEXT,
              created_at TEXT,
              created_by TEXT
            )
        ''')

        db.execute('''
            CREATE TABLE IF NOT EXISTS case_status (
              case_id TEXT PRIMARY KEY,
              stage TEXT,
              percent REAL,
              note TEXT,
              updated_at TEXT,
              updated_by TEXT
            )
        ''')

        db.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              case_id TEXT,
              title TEXT,
              description TEXT,
              priority TEXT DEFAULT 'medium',
              status TEXT DEFAULT 'open',
              owner_id TEXT DEFAULT '',
              due_date TEXT DEFAULT '',
              related_risk_id INTEGER,
              created_at TEXT,
              created_by TEXT,
              updated_at TEXT,
              updated_by TEXT
            )
        ''')
        db.commit()


def ensure_case_status_seed():
    """把 cases table seed 到 case_status（若尚未有記錄）。"""
    try:
        with get_db() as db:
            rows = db.execute('SELECT case_id FROM cases WHERE enabled = 1').fetchall()
            for r in rows:
                cid = r['case_id']
                row = db.execute('SELECT case_id FROM case_status WHERE case_id=?', (cid,)).fetchone()
                if not row:
                    db.execute(
                        'INSERT INTO case_status(case_id, stage, percent, note, updated_at, updated_by) VALUES(?,?,?,?,?,?)',
                        (cid, '', None, '', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system')
                    )
            db.commit()
    except Exception:
        pass


# ===================== 案場清單 =====================
CASES = [
    {"id": "馬偕護專", "name": "馬偕護專停車場", "type": "停車場"},
    {"id": "彰化聯合標租", "name": "彰化學校聯合標租", "type": "屋頂型"},
    {"id": "陸豐國小", "name": "陸豐國小", "type": "屋頂型"},
    {"id": "媽厝國小", "name": "媽厝國小", "type": "屋頂型"},
    {"id": "仁豐國小", "name": "仁豐國小", "type": "屋頂型"},
    {"id": "萬合國小", "name": "萬合國小", "type": "屋頂型"},
    {"id": "長安國小", "name": "長安國小", "type": "屋頂型"},
    {"id": "線西國小", "name": "線西國小H區", "type": "風雨球場"},
    {"id": "伸東國小", "name": "伸東國小", "type": "屋頂型"},
    {"id": "鹿東國小", "name": "鹿東國小二校區", "type": "屋頂型"},
    {"id": "台積電PPA", "name": "台積電 PPA 綠電轉供", "type": "PPA"},
    {"id": "永豐融資", "name": "永豐銀行融資展延", "type": "行政"},
    {"id": "其他", "name": "其他/行政事務", "type": "行政"},
]


def seed_cases_from_constant():
    """把內建 CASES seed 到 cases table（僅補齊，不覆寫管理者自行修改的名稱）。"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        with get_db() as db:
            for c in CASES:
                cid = c.get('id')
                if not cid:
                    continue
                # 若已存在就不覆寫 name/type（讓管理者可在 DB 維護）
                db.execute(
                    "INSERT OR IGNORE INTO cases(case_id, name, type, enabled, created_at, updated_at) VALUES(?,?,?,?,?,?)",
                    (cid, c.get('name', cid), c.get('type', ''), 1, now, now)
                )
            db.commit()
    except Exception:
        pass


def resolve_case_id(raw: str, db=None) -> str:
    """把自由輸入的案場名稱/別名，轉成 canonical case_id。"""
    s = (raw or '').strip()
    if not s:
        return ''

    if db is None:
        try:
            with get_db() as _db:
                return resolve_case_id(s, db=_db)
        except Exception:
            return s

    # 1) alias
    try:
        row = db.execute('SELECT case_id FROM case_aliases WHERE alias = ?', (s,)).fetchone()
        if row and row['case_id']:
            return row['case_id']
    except Exception:
        pass

    # 2) exact case_id exists
    try:
        row = db.execute('SELECT case_id FROM cases WHERE case_id = ? AND enabled = 1', (s,)).fetchone()
        if row and row['case_id']:
            return row['case_id']
    except Exception:
        pass

    return s


# 初始化資料庫
try:
    init_db()
    seed_cases_from_constant()
    ensure_case_status_seed()
except Exception:
    pass

# ===================== 登入系統 =====================
class User(UserMixin):
    def __init__(self, user_data):
        self.id = user_data["id"]
        self.name = user_data["name"]
        self.role = user_data["role"]

@login_manager.user_loader
def load_user(user_id):
    u = USERS.get(user_id)
    if not u:
        return None
    if not bool(u.get('enabled', True)):
        return None
    return User(u)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    # 登入失敗鎖定（避免暴力破解）
    MAX_FAILS = 5
    WINDOW_SEC = 15 * 60
    LOCK_SEC = 15 * 60

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        ip = client_ip()

        key = f"{username}|{ip}"
        st = load_auth_state()
        entry = st.get(key, {'fails': [], 'locked_until': 0})
        now = _now_ts()

        if entry.get('locked_until', 0) > now:
            remain = int(entry['locked_until'] - now)
            mins = max(1, (remain + 59) // 60)
            audit('login_locked', actor_id=username, detail={'mins': mins})
            flash(f'此帳號嘗試次數過多，已暫時鎖定 {mins} 分鐘，請稍後再試或聯繫管理員。', 'error')
            return render_template('login.html')

        user_data = USERS.get(username)

        if user_data and not bool(user_data.get('enabled', True)):
            audit('login_disabled', actor_id=username)
            flash('此帳號已停用，請聯繫管理員。', 'error')
            return render_template('login.html')

        ok = bool(user_data and check_password_hash(user_data.get('password_hash', ''), password))

        if ok:
            st[key] = {'fails': [], 'locked_until': 0}
            save_auth_state(st)

            user = User(user_data)
            login_user(user)
            session.permanent = True
            audit('login_success', actor_id=user.id)
            flash(f'歡迎回來，{user.name}！', 'success')
            return redirect(url_for('index'))

        fails = [t for t in entry.get('fails', []) if isinstance(t, int) and now - t <= WINDOW_SEC]
        fails.append(now)
        locked_until = 0
        if len(fails) >= MAX_FAILS:
            locked_until = now + LOCK_SEC
        st[key] = {'fails': fails, 'locked_until': locked_until}
        save_auth_state(st)
        audit('login_fail', actor_id=username, detail={'fail_count': len(fails), 'locked': bool(locked_until)})
        flash('帳號或密碼錯誤', 'error')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    audit('logout', actor_id=current_user.id if current_user.is_authenticated else '')
    logout_user()
    flash('已登出', 'info')
    return redirect(url_for('login'))


@app.before_request
def force_password_change():
    """若帳號仍為臨時密碼，強制導向修改密碼頁。"""
    try:
        if current_user.is_authenticated:
            u = USERS.get(current_user.id)
            if u and u.get('password_temp'):
                # 允許的 endpoint
                allow = {'change_password', 'logout', 'static', 'login'}
                if request.endpoint not in allow and not (request.path or '').startswith('/static'):
                    return redirect(url_for('change_password'))
    except Exception:
        pass


def _validate_new_password(pw: str):
    """基本密碼規則：至少 8 碼，且同時包含英文字母與數字。"""
    if not pw or len(pw) < 8:
        return False, '密碼長度需至少 8 碼'
    has_alpha = any(c.isalpha() for c in pw)
    has_digit = any(c.isdigit() for c in pw)
    if not (has_alpha and has_digit):
        return False, '密碼需同時包含英文字母與數字'
    if ' ' in pw:
        return False, '密碼不可包含空白'
    return True, ''


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    u = USERS.get(current_user.id)
    if not u:
        flash('使用者不存在', 'error')
        return redirect(url_for('logout'))

    forced = bool(u.get('password_temp'))

    if request.method == 'POST':
        old_pw = request.form.get('old_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        if not check_password_hash(u.get('password_hash', ''), old_pw):
            flash('舊密碼不正確', 'error')
            return redirect(url_for('change_password'))
        if new_pw != confirm_pw:
            flash('新密碼與確認密碼不一致', 'error')
            return redirect(url_for('change_password'))
        ok, msg = _validate_new_password(new_pw)
        if not ok:
            flash(msg, 'error')
            return redirect(url_for('change_password'))

        u['password_hash'] = generate_password_hash(new_pw)
        u['password_temp'] = False
        USERS[current_user.id] = u
        save_users(USERS)
        refresh_employee_cache()
        audit('password_change', actor_id=current_user.id)
        flash('✅ 密碼已更新', 'success')
        return redirect(url_for('index'))

    return render_template('change_password.html', forced=forced)

# ===================== 首頁 =====================
@app.route('/')
@login_required
def index():
    today = datetime.now().strftime('%Y-%m-%d')
    today_dir = REPORTS_DIR / today

    submitted_ids = set()
    if today_dir.exists():
        for f in today_dir.glob('*_report.json'):
            submitted_ids.add(f.stem.replace('_report', ''))

    submitted = [e for e in EMPLOYEES if e['id'] in submitted_ids]
    not_submitted = [e for e in EMPLOYEES if e['id'] not in submitted_ids]
    total = len(EMPLOYEES)
    submitted_count = len(submitted)
    rate = round(submitted_count / total * 100, 1) if total else 0

    # 員工只能看到自己的按鈕
    is_admin = current_user.role == 'admin'

    risk_top5 = []
    overdue_tasks_top5 = []
    if is_admin:
        try:
            with get_db() as db:
                risk_top5 = db.execute(
                    """
                    SELECT id, case_id, category, level, risk_desc, status, owner_id, due_date, report_date, employee_id
                    FROM risk_items
                    WHERE status != 'closed' AND level = 'high'
                    ORDER BY
                      CASE WHEN due_date != '' AND due_date < ? THEN 0
                           WHEN due_date != '' THEN 1
                           ELSE 2 END,
                      due_date ASC, id DESC
                    LIMIT 5
                    """,
                    (today,)
                ).fetchall()

                overdue_tasks_top5 = db.execute(
                    """
                    SELECT id, case_id, title, priority, status, owner_id, due_date, related_risk_id
                    FROM tasks
                    WHERE status != 'closed' AND due_date != '' AND due_date < ?
                    ORDER BY due_date ASC, id ASC
                    LIMIT 5
                    """,
                    (today,)
                ).fetchall()
        except Exception:
            pass

    return render_template('index_v3.html',
        employees=EMPLOYEES,
        submitted=submitted,
        not_submitted=not_submitted,
        total=total,
        submitted_count=submitted_count,
        submission_rate=rate,
        today=today,
        is_admin=is_admin,
        current_user_id=current_user.id,
        risk_top5=risk_top5,
        overdue_tasks_top5=overdue_tasks_top5,
    )

def safe_float(v, default=0.0):
    try:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == '':
            return default
        return float(s)
    except Exception:
        return default


def normalize_report_items(work_items, plan_items, risk_items):
    """清理空資料、修正欄位缺失，避免資料庫被垃圾資料污染。"""
    db = None
    try:
        db = get_db()
    except Exception:
        db = None

    # work: 至少要有案場 + 工作內容
    clean_work = []
    for w in (work_items or []):
        case_id = resolve_case_id((w.get('case_id') or '').strip(), db=db)
        content = (w.get('work_content') or '').strip()
        if not case_id or not content:
            continue
        clean_work.append({
            'case_id': case_id,
            'work_type': (w.get('work_type') or '').strip(),
            'progress': min(100.0, max(0.0, safe_float(w.get('progress'), 0.0))),
            'hours': max(0.0, safe_float(w.get('hours'), 0.0)),
            'work_content': content,
            'output': (w.get('output') or '').strip(),
            'status': (w.get('status') or '').strip(),
        })

    # plan: 至少要有案場 + 計畫內容
    clean_plan = []
    for p in (plan_items or []):
        case_id = resolve_case_id((p.get('case_id') or '').strip(), db=db)
        content = (p.get('plan_content') or '').strip()
        if not case_id or not content:
            continue
        clean_plan.append({
            'case_id': case_id,
            'plan_content': content,
            'plan_hours': max(0.0, safe_float(p.get('plan_hours'), 0.0)),
            'need_support': (p.get('need_support') or '').strip(),
        })

    # risk: 至少要有案場 + 風險說明
    clean_risk = []
    for r in (risk_items or []):
        case_id = resolve_case_id((r.get('case_id') or '').strip(), db=db)
        desc = (r.get('risk_desc') or '').strip()
        if not case_id or not desc:
            continue
        cat = (r.get('category') or '協調').strip()
        if cat not in RISK_CATEGORIES:
            cat = '其他'
        lv = (r.get('risk_level') or 'low').strip()
        if lv not in ('high','medium','low'):
            lv = 'low'
        clean_risk.append({
            'case_id': case_id,
            'category': cat,
            'risk_level': lv,
            'risk_desc': desc,
            'risk_impact': (r.get('risk_impact') or '').strip(),
            'need_help': (r.get('need_help') or '').strip(),
        })

    try:
        if db:
            db.close()
    except Exception:
        pass

    return clean_work, clean_plan, clean_risk


ALLOWED_EXTENSIONS = {
    'jpg','jpeg','png','gif','webp',
    'pdf','txt',
    'doc','docx',
    'xls','xlsx',
    'ppt','pptx',
    'zip'
}


def allowed_file(filename: str) -> bool:
    if not filename:
        return False
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS


def save_uploaded_files(employee_id: str, report_date: str, submit_time: str, kind: str, files) -> list[dict]:
    """儲存上傳附件到磁碟，回傳檔案 metadata 清單。"""
    saved = []
    if not files:
        return saved

    year = report_date[:4]
    # 依 年份 / 日期 / 人員 / 提交時間 分類，方便日後歸檔備份
    batch = submit_time.replace('-', '').replace(':', '').replace(' ', '_')  # YYYYMMDD_HHMMSS
    base_dir = ATTACHMENTS_DIR / year / report_date / employee_id / batch / kind
    base_dir.mkdir(parents=True, exist_ok=True)

    for f in files:
        if not f or not getattr(f, 'filename', ''):
            continue
        orig = f.filename
        if not allowed_file(orig):
            continue
        safe = secure_filename(orig)
        # 再加一層隨機字串避免撞名
        token = secrets.token_hex(4)
        stored = f"{token}__{safe}" if safe else f"{token}"
        out_path = base_dir / stored
        f.save(out_path)

        rel = str(out_path.relative_to(BASE_DIR))
        try:
            size = out_path.stat().st_size
        except Exception:
            size = None

        saved.append({
            'kind': kind,
            'original_name': orig,
            'stored_name': stored,
            'rel_path': rel,
            'size_bytes': size,
        })

    return saved


# ===================== 填寫表單 =====================
@app.route('/report', methods=['GET', 'POST'])
@app.route('/report/<employee_id>', methods=['GET', 'POST'])
@login_required
def report_form(employee_id=None):
    # 員工只能填自己的日報
    if current_user.role == 'employee':
        employee_id = current_user.id

    user_data = USERS.get(employee_id)
    if not user_data:
        flash('員工不存在', 'error')
        return redirect(url_for('index'))

    today = datetime.now().strftime('%Y-%m-%d')

    # 檢查是否已提交（允許覆蓋）
    today_dir = REPORTS_DIR / today
    existing = None
    report_file = today_dir / f"{employee_id}_report.json"
    if report_file.exists():
        with open(report_file, encoding='utf-8') as f:
            existing = json.load(f)

    if request.method == 'POST':
        data = request.json if request.is_json else request.form

        work_items = json.loads(request.form.get('work_items_json', '[]'))
        plan_items = json.loads(request.form.get('plan_items_json', '[]'))
        risk_items = json.loads(request.form.get('risk_items_json', '[]'))

        # 清理與正規化（避免空白資料/不合法值）
        work_items, plan_items, risk_items = normalize_report_items(work_items, plan_items, risk_items)

        if not work_items:
            flash('請至少填寫 1 筆「今日工作」（需包含案場與工作內容）', 'error')
            return redirect(request.url)

        submit_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 附件上傳（依 年份/日期/人員/提交時間/類型 分類儲存）
        uploaded_files = []
        try:
            uploaded_files += save_uploaded_files(employee_id, today, submit_time, 'photo', request.files.getlist('files_photo'))
            uploaded_files += save_uploaded_files(employee_id, today, submit_time, 'meeting', request.files.getlist('files_meeting'))
            uploaded_files += save_uploaded_files(employee_id, today, submit_time, 'document', request.files.getlist('files_document'))
            uploaded_files += save_uploaded_files(employee_id, today, submit_time, 'other', request.files.getlist('files_other'))
        except Exception:
            # 附件失敗不應阻擋日報主流程
            uploaded_files = []

        report = {
            "employee_id": employee_id,
            "employee_name": user_data['name'],
            "department": "未分類",
            "report_date": today,
            "submit_time": submit_time,
            "work_items": work_items,
            "plan_items": plan_items,
            "risk_items": risk_items,
            "summary": {
                "today_gain": request.form.get('today_gain', ''),
                "improvement": request.form.get('improvement', ''),
                "remarks": request.form.get('remarks', '')
            },
            "attachments": {
                "photo": any(f.get('kind')=='photo' for f in uploaded_files),
                "meeting": any(f.get('kind')=='meeting' for f in uploaded_files),
                "document": any(f.get('kind')=='document' for f in uploaded_files),
                "other": any(f.get('kind')=='other' for f in uploaded_files),
                "files": uploaded_files,
                "note": request.form.get('attachment_note', '')
            }
        }

        today_dir.mkdir(parents=True, exist_ok=True)
        is_edit = report_file.exists()
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        # 同步寫入 SQLite（供後續案件儀表板/風險追蹤/AI 取用）
        try:
            report_key = f"{today}|{employee_id}"
            with get_db() as db:
                db.execute('REPLACE INTO reports(report_key, report_date, employee_id, submitted_at, report_json) VALUES(?,?,?,?,?)',
                           (report_key, today, employee_id, report['submit_time'], json.dumps(report, ensure_ascii=False)))
                db.execute('DELETE FROM work_items WHERE report_key=?', (report_key,))
                db.execute('DELETE FROM plan_items WHERE report_key=?', (report_key,))
                db.execute('DELETE FROM risk_items WHERE report_key=?', (report_key,))

                for w in work_items:
                    db.execute(
                        'INSERT INTO work_items(report_key, case_id, work_type, progress, hours, work_content, output, status) VALUES(?,?,?,?,?,?,?,?)',
                        (
                            report_key,
                            (w.get('case_id') or '').strip(),
                            (w.get('work_type') or '').strip(),
                            safe_float(w.get('progress'), 0.0),
                            safe_float(w.get('hours'), 0.0),
                            (w.get('work_content') or '').strip(),
                            (w.get('output') or '').strip(),
                            (w.get('status') or '').strip(),
                        )
                    )

                
                for p in plan_items:
                    db.execute(
                        'INSERT INTO plan_items(report_key, case_id, plan_content, plan_hours, need_support) VALUES(?,?,?,?,?)',
                        (
                            report_key,
                            (p.get('case_id') or '').strip(),
                            (p.get('plan_content') or '').strip(),
                            safe_float(p.get('plan_hours'), 0.0),
                            (p.get('need_support') or '').strip(),
                        )
                    )

                for r in risk_items:
                    db.execute(
                        'INSERT INTO risk_items(report_key, report_date, employee_id, case_id, category, level, risk_desc, risk_impact, need_help, created_at) VALUES(?,?,?,?,?,?,?,?,?,?)',
                        (
                            report_key,
                            today,
                            employee_id,
                            (r.get('case_id') or '').strip(),
                            (r.get('category') or '協調').strip(),
                            (r.get('risk_level') or 'low').strip(),
                            (r.get('risk_desc') or '').strip(),
                            (r.get('risk_impact') or '').strip(),
                            (r.get('need_help') or '').strip(),
                            report['submit_time'],
                        )
                    )

                # 附件 metadata（不刪舊檔，以提交時間分批保存）
                for a in uploaded_files:
                    db.execute(
                        'INSERT INTO attachments(report_key, report_date, employee_id, kind, original_name, stored_name, rel_path, size_bytes, uploaded_at) VALUES(?,?,?,?,?,?,?,?,?)',
                        (
                            report_key,
                            today,
                            employee_id,
                            a.get('kind',''),
                            a.get('original_name',''),
                            a.get('stored_name',''),
                            a.get('rel_path',''),
                            a.get('size_bytes'),
                            report['submit_time'],
                        )
                    )

                db.commit()
        except Exception:
            pass

        audit('report_submit', actor_id=current_user.id, target_id=employee_id, detail={'date': today, 'edit': bool(is_edit)})

        # 編輯留痕
        action = '覆蓋修改' if is_edit else '初次提交'
        append_edit_log(report_file, current_user.name, action)

        # 自動觸發 AI 分析（背景執行）
        def run_analysis():
            try:
                subprocess.run(
                    ['python3', str(SCRIPTS_DIR / 'analyze_daily_reports.py'), today],
                    capture_output=True, timeout=60
                )
            except Exception:
                pass
        threading.Thread(target=run_analysis, daemon=True).start()

        flash('✅ 日報提交成功！AI 分析已自動觸發。', 'success')
        return redirect(url_for('index'))

    return render_template('report_form_v3.html',
        employee=user_data,
        today=today,
        cases=get_case_catalog(),
        existing=existing
    )

# ===================== 草稿 API =====================
@app.route('/api/draft/save', methods=['POST'])
@login_required
def save_draft():
    today = datetime.now().strftime('%Y-%m-%d')
    draft_dir = REPORTS_DIR / today
    draft_dir.mkdir(parents=True, exist_ok=True)
    draft_file = draft_dir / f"{current_user.id}_draft.json"
    with open(draft_file, 'w', encoding='utf-8') as f:
        json.dump(request.json, f, ensure_ascii=False, indent=2)
    return jsonify({"status": "saved"})

@app.route('/api/draft/load')
@login_required
def load_draft():
    today = datetime.now().strftime('%Y-%m-%d')
    draft_file = REPORTS_DIR / today / f"{current_user.id}_draft.json"
    if draft_file.exists():
        with open(draft_file, encoding='utf-8') as f:
            return jsonify(json.load(f))
    return jsonify({})

# ===================== 案場字典（管理員） =====================
@app.route('/admin/cases')
@login_required
def admin_cases():
    if current_user.role != 'admin':
        flash('無權限訪問', 'error')
        return redirect(url_for('index'))

    with get_db() as db:
        canonical = db.execute('SELECT case_id, name, type, enabled FROM cases ORDER BY enabled DESC, case_id ASC').fetchall()
        aliases = db.execute('SELECT alias, case_id, created_at, created_by FROM case_aliases ORDER BY alias ASC').fetchall()

        # 找出 DB 中出現過但尚未登錄在 cases table 的案場（通常是自由輸入或舊資料）
        unknown = db.execute(
            """
            SELECT case_id FROM (
              SELECT DISTINCT case_id FROM work_items WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM risk_items WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM tasks WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM plan_items WHERE case_id!=''
            )
            WHERE case_id NOT IN (SELECT case_id FROM cases)
            ORDER BY case_id ASC
            """
        ).fetchall()

    return render_template('admin_cases.html', canonical=canonical, aliases=aliases, unknown=unknown)


@app.route('/admin/cases/create', methods=['POST'])
@login_required
def admin_cases_create():
    if current_user.role != 'admin':
        abort(403)

    case_id = (request.form.get('case_id') or '').strip()
    name = (request.form.get('name') or '').strip()
    ctype = (request.form.get('type') or '').strip()

    if not case_id or not name:
        flash('案場代碼與名稱不可空白', 'error')
        return redirect(url_for('admin_cases'))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        exists = db.execute('SELECT case_id FROM cases WHERE case_id=?', (case_id,)).fetchone()
        if exists:
            flash('案場代碼已存在', 'error')
            return redirect(url_for('admin_cases'))

        db.execute('INSERT INTO cases(case_id, name, type, enabled, created_at, updated_at) VALUES(?,?,?,?,?,?)',
                   (case_id, name, ctype, 1, now, now))
        db.execute('INSERT OR IGNORE INTO case_status(case_id, stage, percent, note, updated_at, updated_by) VALUES(?,?,?,?,?,?)',
                   (case_id, '', None, '', now, current_user.id))
        db.commit()

    audit('case_create', actor_id=current_user.id, target_id=case_id)
    flash('✅ 案場已新增', 'success')
    return redirect(url_for('admin_cases'))


@app.route('/admin/cases/update/<case_id>', methods=['POST'])
@login_required
def admin_cases_update(case_id):
    if current_user.role != 'admin':
        abort(403)

    name = (request.form.get('name') or '').strip()
    ctype = (request.form.get('type') or '').strip()
    enabled = request.form.get('enabled') == 'on'

    if not name:
        flash('案場名稱不可空白', 'error')
        return redirect(url_for('admin_cases'))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        db.execute('UPDATE cases SET name=?, type=?, enabled=?, updated_at=? WHERE case_id=?',
                   (name, ctype, 1 if enabled else 0, now, case_id))
        db.commit()

    audit('case_update', actor_id=current_user.id, target_id=case_id, detail={'enabled': enabled})
    flash('✅ 案場已更新', 'success')
    return redirect(url_for('admin_cases'))


@app.route('/admin/cases/alias/add', methods=['POST'])
@login_required
def admin_case_alias_add():
    if current_user.role != 'admin':
        abort(403)

    alias = (request.form.get('alias') or '').strip()
    case_id = (request.form.get('case_id') or '').strip()
    if not alias or not case_id:
        flash('別名與對應案場不可空白', 'error')
        return redirect(url_for('admin_cases'))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        ok = db.execute('SELECT case_id FROM cases WHERE case_id=?', (case_id,)).fetchone()
        if not ok:
            flash('對應的案場不存在，請先新增案場', 'error')
            return redirect(url_for('admin_cases'))

        # alias 不允許覆蓋 canonical case_id 本身（避免混淆）
        exists = db.execute('SELECT case_id FROM cases WHERE case_id=?', (alias,)).fetchone()
        if exists:
            flash('此別名與既有案場代碼相同，請改用其他別名', 'error')
            return redirect(url_for('admin_cases'))

        db.execute('REPLACE INTO case_aliases(alias, case_id, created_at, created_by) VALUES(?,?,?,?)',
                   (alias, case_id, now, current_user.id))
        db.commit()

    audit('case_alias_add', actor_id=current_user.id, target_id=alias, detail={'case_id': case_id})
    flash('✅ 別名已新增/更新', 'success')
    return redirect(url_for('admin_cases'))


@app.route('/admin/cases/alias/delete/<alias>', methods=['POST'])
@login_required
def admin_case_alias_delete(alias):
    if current_user.role != 'admin':
        abort(403)

    with get_db() as db:
        db.execute('DELETE FROM case_aliases WHERE alias=?', (alias,))
        db.commit()

    audit('case_alias_delete', actor_id=current_user.id, target_id=alias)
    flash('✅ 別名已刪除', 'success')
    return redirect(url_for('admin_cases'))


# ===================== 帳號/權限管理（管理員） =====================
@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        flash('無權限訪問', 'error')
        return redirect(url_for('index'))

    # 排序：admin 先，其次啟用，再依 id
    rows = list(USERS.values())
    rows.sort(key=lambda u: (0 if u.get('role')=='admin' else 1, 0 if bool(u.get('enabled',True)) else 1, u.get('id','')))

    return render_template('admin_users.html', users=rows)


def _is_valid_user_id(uid: str) -> bool:
    # 允許英數字底線（方便做 Owner / API）
    return bool(re.fullmatch(r'[a-zA-Z0-9_]{3,64}', uid or ''))


def _count_enabled_admins() -> int:
    return sum(1 for u in USERS.values() if u.get('role')=='admin' and bool(u.get('enabled', True)))


@app.route('/admin/users/create', methods=['POST'])
@login_required
def admin_users_create():
    if current_user.role != 'admin':
        abort(403)

    uid = (request.form.get('id') or '').strip()
    name = (request.form.get('name') or '').strip()
    role = (request.form.get('role') or 'employee').strip()

    if not _is_valid_user_id(uid):
        flash('帳號 ID 格式不正確（需 3~64 碼英數字或底線）', 'error')
        return redirect(url_for('admin_users'))
    if uid in USERS:
        flash('帳號 ID 已存在', 'error')
        return redirect(url_for('admin_users'))
    if not name:
        flash('姓名不可空白', 'error')
        return redirect(url_for('admin_users'))
    if role not in ('admin', 'employee'):
        role = 'employee'

    temp_pw = secrets.token_urlsafe(9)
    USERS[uid] = {
        'id': uid,
        'name': name,
        'role': role,
        'enabled': True,
        'password_hash': generate_password_hash(temp_pw),
        'password_temp': True,
    }
    save_users(USERS)
    refresh_employee_cache()

    audit('user_create', actor_id=current_user.id, target_id=uid, detail={'role': role})
    flash(f'✅ 已建立帳號 {uid}（{name}）。臨時密碼：{temp_pw}（請轉交本人並要求首次登入立即修改）', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/update/<user_id>', methods=['POST'])
@login_required
def admin_users_update(user_id):
    if current_user.role != 'admin':
        abort(403)

    u = USERS.get(user_id)
    if not u:
        flash('使用者不存在', 'error')
        return redirect(url_for('admin_users'))

    name = (request.form.get('name') or u.get('name') or '').strip()
    role = (request.form.get('role') or u.get('role') or 'employee').strip()
    enabled = request.form.get('enabled') == 'on'

    # 至少要保留 1 個啟用的 admin
    if u.get('role') == 'admin' and (not enabled or role != 'admin'):
        if _count_enabled_admins() <= 1 and bool(u.get('enabled', True)):
            flash('至少需保留 1 位啟用的管理員，無法停用/降權最後一位管理員', 'error')
            return redirect(url_for('admin_users'))

    # 不允許把自己停用（避免鎖死）
    if user_id == current_user.id and not enabled:
        flash('不可停用自己', 'error')
        return redirect(url_for('admin_users'))

    if role not in ('admin', 'employee'):
        role = u.get('role')

    u['name'] = name
    u['role'] = role
    u['enabled'] = enabled
    USERS[user_id] = u
    save_users(USERS)
    refresh_employee_cache()

    audit('user_update', actor_id=current_user.id, target_id=user_id, detail={'role': role, 'enabled': enabled})
    flash('✅ 使用者已更新', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/delete/<user_id>', methods=['POST'])
@login_required
def admin_users_delete(user_id):
    if current_user.role != 'admin':
        abort(403)

    if user_id == current_user.id:
        flash('不可刪除自己', 'error')
        return redirect(url_for('admin_users'))

    u = USERS.get(user_id)
    if not u:
        flash('使用者不存在', 'error')
        return redirect(url_for('admin_users'))

    if u.get('role') == 'admin' and bool(u.get('enabled', True)) and _count_enabled_admins() <= 1:
        flash('至少需保留 1 位啟用的管理員，無法刪除最後一位管理員', 'error')
        return redirect(url_for('admin_users'))

    USERS.pop(user_id, None)
    save_users(USERS)
    refresh_employee_cache()

    audit('user_delete', actor_id=current_user.id, target_id=user_id)
    flash('✅ 使用者已刪除', 'success')
    return redirect(url_for('admin_users'))


# ===================== 稽核紀錄（管理員） =====================
@app.route('/admin/audit')
@login_required
def admin_audit():
    if current_user.role != 'admin':
        flash('無權限訪問', 'error')
        return redirect(url_for('index'))

    q_event = request.args.get('event', '').strip()
    q_actor = request.args.get('actor', '').strip()

    rows = tail_jsonl(AUDIT_LOG_FILE, max_lines=500)
    rows.sort(key=lambda r: r.get('ts', 0), reverse=True)

    def ok(r):
        if q_event and r.get('event') != q_event:
            return False
        if q_actor and r.get('actor_id') != q_actor:
            return False
        return True

    filtered = [r for r in rows if ok(r)][:300]

    events = sorted({r.get('event') for r in rows if r.get('event')})
    actors = sorted({r.get('actor_id') for r in rows if r.get('actor_id')})

    return render_template('admin_audit.html', rows=filtered, events=events, actors=actors, q_event=q_event, q_actor=q_actor)


# ===================== 管理後台 =====================
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('無權限訪問', 'error')
        return redirect(url_for('index'))

    today = datetime.now().strftime('%Y-%m-%d')
    date_param = request.args.get('date', today)
    date_dir = REPORTS_DIR / date_param

    reports = []
    if date_dir.exists():
        for f in sorted(date_dir.glob('*_report.json')):
            with open(f, encoding='utf-8') as fp:
                reports.append(json.load(fp))

    analysis = None
    analysis_file = date_dir / 'analysis_result.json'
    if analysis_file.exists():
        with open(analysis_file, encoding='utf-8') as f:
            analysis = json.load(f)

    # 取得歷史日期清單
    history_dates = sorted(
        [d.name for d in REPORTS_DIR.iterdir() if d.is_dir()],
        reverse=True
    )[:30]

    submitted_ids = {r['employee_id'] for r in reports}
    not_submitted = [e for e in EMPLOYEES if e['id'] not in submitted_ids]

    return render_template('admin_v3.html',
        reports=reports,
        analysis=analysis,
        date=date_param,
        today=today,
        history_dates=history_dates,
        employees=EMPLOYEES,
        not_submitted=not_submitted,
        submitted_count=len(reports),
        total=len(EMPLOYEES)
    )

# ===================== 歷史查詢 =====================
@app.route('/history')
@login_required
def history():
    emp_id = current_user.id if current_user.role == 'employee' else request.args.get('emp_id', '')
    records = []
    if REPORTS_DIR.exists():
        for date_dir in sorted(REPORTS_DIR.iterdir(), reverse=True):
            if not date_dir.is_dir():
                continue
            report_file = date_dir / f"{emp_id}_report.json"
            if report_file.exists():
                with open(report_file, encoding='utf-8') as f:
                    records.append(json.load(f))
    return render_template('history.html',
        records=records,
        emp_id=emp_id,
        employees=EMPLOYEES,
        is_admin=current_user.role == 'admin'
    )

# ===================== 案場 / 風險（案件中心管理） =====================

def get_case_catalog(db=None):
    """回傳案場清單（canonical cases + DB 中實際出現過但未登錄的案場）。"""

    if db is None:
        try:
            with get_db() as _db:
                return get_case_catalog(_db)
        except Exception:
            # fallback: 使用內建 CASES
            return [{'id': c['id'], 'name': c.get('name', c['id']), 'type': c.get('type', '')} for c in CASES]

    base = {}

    # 1) canonical cases
    try:
        rows = db.execute('SELECT case_id, name, type FROM cases WHERE enabled = 1').fetchall()
        for r in rows:
            base[r['case_id']] = {'id': r['case_id'], 'name': r['name'] or r['case_id'], 'type': r['type'] or ''}
    except Exception:
        for c in CASES:
            base[c['id']] = {'id': c['id'], 'name': c.get('name', c['id']), 'type': c.get('type', '')}

    # 2) 追加 DB 中出現過但未登錄的案場（標記未登錄，方便管理者收斂）
    try:
        rows = db.execute(
            """
            SELECT case_id FROM (
              SELECT DISTINCT case_id FROM work_items WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM risk_items WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM tasks WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM plan_items WHERE case_id!=''
            )
            """
        ).fetchall()
        for r in rows:
            cid = r['case_id']
            if cid and cid not in base:
                base[cid] = {'id': cid, 'name': cid, 'type': '未登錄'}
    except Exception:
        pass

    return list(base.values())

    try:
        rows = db.execute(
            """
            SELECT case_id FROM (
              SELECT DISTINCT case_id FROM work_items WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM risk_items WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM tasks WHERE case_id!=''
              UNION
              SELECT DISTINCT case_id FROM plan_items WHERE case_id!=''
            )
            """
        ).fetchall()
        for r in rows:
            cid = r['case_id'] if isinstance(r, dict) else r[0]
            if cid and cid not in base:
                base[cid] = {'id': cid, 'name': cid, 'type': '未登錄'}
    except Exception:
        pass

    return list(base.values())


def case_by_id(case_id: str, db=None):
    for c in get_case_catalog(db=db):
        if c.get('id') == case_id:
            return c
    return {'id': case_id, 'name': case_id, 'type': '未登錄'}


@app.route('/cases')
@login_required
def cases_dashboard():
    today = datetime.now().strftime('%Y-%m-%d')

    with get_db() as db:
        catalog = get_case_catalog(db)

        # 員工只看到自己碰過的案場；管理員看到全部
        visible = catalog
        if current_user.role == 'employee':
            rows = db.execute(
                """
                SELECT DISTINCT w.case_id AS case_id
                FROM work_items w
                JOIN reports r ON w.report_key = r.report_key
                WHERE r.employee_id = ? AND w.case_id != ''
                """,
                (current_user.id,)
            ).fetchall()
            touched = {r['case_id'] for r in rows}

            rows2 = db.execute(
                "SELECT DISTINCT case_id FROM risk_items WHERE employee_id=? AND case_id!=''",
                (current_user.id,)
            ).fetchall()
            touched |= {r['case_id'] for r in rows2}

            rows3 = db.execute(
                "SELECT DISTINCT case_id FROM tasks WHERE owner_id=? AND case_id!=''",
                (current_user.id,)
            ).fetchall()
            touched |= {r['case_id'] for r in rows3}

            visible = [c for c in catalog if c['id'] in touched]

        # 補齊 case_status（若 DB 新出現了未登錄案場，也要能被管理）
        for c in visible:
            db.execute(
                'INSERT OR IGNORE INTO case_status(case_id, stage, percent, note, updated_at, updated_by) VALUES(?,?,?,?,?,?)',
                (c['id'], '', None, '', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system')
            )
        db.commit()

        status_rows = db.execute('SELECT * FROM case_status').fetchall()
        status_map = {r['case_id']: r for r in status_rows}

        summaries = []
        for c in visible:
            cid = c['id']
            st = status_map.get(cid)

            r1 = db.execute(
                """
                SELECT MAX(r.report_date) AS last_date
                FROM work_items w
                JOIN reports r ON w.report_key = r.report_key
                WHERE w.case_id = ?
                """,
                (cid,)
            ).fetchone()
            r2 = db.execute(
                "SELECT MAX(report_date) AS last_date FROM risk_items WHERE case_id = ?",
                (cid,)
            ).fetchone()
            r3 = db.execute(
                "SELECT MAX(due_date) AS last_due FROM tasks WHERE case_id = ?",
                (cid,)
            ).fetchone()
            last_date = max((r1['last_date'] or ''), (r2['last_date'] or ''))

            rc = db.execute(
                """
                SELECT
                  SUM(CASE WHEN status != 'closed' THEN 1 ELSE 0 END) AS total,
                  SUM(CASE WHEN status != 'closed' AND level = 'high' THEN 1 ELSE 0 END) AS high,
                  SUM(CASE WHEN status != 'closed' AND level = 'medium' THEN 1 ELSE 0 END) AS medium,
                  SUM(CASE WHEN status != 'closed' AND level = 'low' THEN 1 ELSE 0 END) AS low
                FROM risk_items
                WHERE case_id = ?
                """,
                (cid,)
            ).fetchone()

            tc = db.execute(
                """
                SELECT
                  SUM(CASE WHEN status != 'closed' THEN 1 ELSE 0 END) AS open_total,
                  SUM(CASE WHEN status != 'closed' AND due_date != '' AND due_date < ? THEN 1 ELSE 0 END) AS overdue
                FROM tasks
                WHERE case_id = ?
                """,
                (today, cid)
            ).fetchone()

            summaries.append({
                'id': cid,
                'name': c.get('name', cid),
                'type': c.get('type', ''),
                'stage': (st['stage'] if st else ''),
                'percent': (st['percent'] if st and st['percent'] is not None else ''),
                'updated_at': (st['updated_at'] if st else ''),
                'updated_by': (st['updated_by'] if st else ''),
                'last_date': last_date,
                'risk_total': rc['total'] or 0,
                'risk_high': rc['high'] or 0,
                'risk_medium': rc['medium'] or 0,
                'risk_low': rc['low'] or 0,
                'task_open_total': tc['open_total'] or 0,
                'task_overdue': tc['overdue'] or 0,
            })

    # 預設排序：高風險多 → 逾期多 → 未結案風險總量
    summaries.sort(key=lambda x: (x['risk_high'], x['task_overdue'], x['risk_total']), reverse=True)

    return render_template('cases.html', cases=summaries, is_admin=current_user.role == 'admin')


@app.route('/cases/<case_id>', methods=['GET', 'POST'])
@login_required
def case_detail(case_id):
    with get_db() as db:
        case = case_by_id(case_id, db=db)

        # 確保此案場在 case_status 內可被管理
        db.execute(
            'INSERT OR IGNORE INTO case_status(case_id, stage, percent, note, updated_at, updated_by) VALUES(?,?,?,?,?,?)',
            (case_id, '', None, '', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system')
        )
        db.commit()
        if request.method == 'POST':
            if current_user.role != 'admin':
                abort(403)
            stage = (request.form.get('stage') or '').strip()
            percent_raw = (request.form.get('percent') or '').strip()
            note = (request.form.get('note') or '').strip()

            if stage and stage not in MILESTONE_STAGES:
                flash('里程碑階段不正確', 'error')
                return redirect(url_for('case_detail', case_id=case_id))

            percent = None
            if percent_raw != '':
                try:
                    percent = float(percent_raw)
                except Exception:
                    flash('百分比需為數字', 'error')
                    return redirect(url_for('case_detail', case_id=case_id))
                if percent < 0 or percent > 100:
                    flash('百分比需在 0~100 之間', 'error')
                    return redirect(url_for('case_detail', case_id=case_id))

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db.execute(
                'REPLACE INTO case_status(case_id, stage, percent, note, updated_at, updated_by) VALUES(?,?,?,?,?,?)',
                (case_id, stage, percent, note, now, current_user.id)
            )
            db.commit()
            audit('case_status_update', actor_id=current_user.id, target_id=case_id, detail={'stage': stage, 'percent': percent})
            flash('✅ 案場進度已更新', 'success')
            return redirect(url_for('case_detail', case_id=case_id))

        st = db.execute('SELECT * FROM case_status WHERE case_id = ?', (case_id,)).fetchone()

        # 近期工作明細（最近 60 筆）
        works = db.execute(
            """
            SELECT r.report_date, r.employee_id, w.*
            FROM work_items w
            JOIN reports r ON w.report_key = r.report_key
            WHERE w.case_id = ?
            ORDER BY r.report_date DESC, w.id DESC
            LIMIT 60
            """,
            (case_id,)
        ).fetchall()

        # 風險清單（最近 80 筆）
        if current_user.role == 'admin':
            risks = db.execute(
                "SELECT * FROM risk_items WHERE case_id = ? ORDER BY id DESC LIMIT 80",
                (case_id,)
            ).fetchall()
        else:
            risks = db.execute(
                "SELECT * FROM risk_items WHERE case_id = ? AND employee_id = ? ORDER BY id DESC LIMIT 80",
                (case_id, current_user.id)
            ).fetchall()

        # 明日計畫（最近 60 筆）
        plans = db.execute(
            """
            SELECT r.report_date, r.employee_id, p.*
            FROM plan_items p
            JOIN reports r ON p.report_key = r.report_key
            WHERE p.case_id = ?
            ORDER BY r.report_date DESC, p.id DESC
            LIMIT 60
            """,
            (case_id,)
        ).fetchall()

        # 任務清單（案件中心追蹤）
        if current_user.role == 'admin':
            tasks = db.execute(
                "SELECT * FROM tasks WHERE case_id = ? ORDER BY CASE status WHEN 'open' THEN 0 WHEN 'in_progress' THEN 1 ELSE 2 END, due_date DESC, id DESC LIMIT 200",
                (case_id,)
            ).fetchall()
        else:
            tasks = db.execute(
                "SELECT * FROM tasks WHERE case_id = ? AND owner_id = ? ORDER BY CASE status WHEN 'open' THEN 0 WHEN 'in_progress' THEN 1 ELSE 2 END, due_date DESC, id DESC LIMIT 200",
                (case_id, current_user.id)
            ).fetchall()

        owners = [{'id': 'admin', 'name': '管理員'}] + list(EMPLOYEES)

    return render_template(
        'case_detail.html',
        case=case,
        status=st,
        works=works,
        plans=plans,
        risks=risks,
        tasks=tasks,
        owners=owners,
        is_admin=current_user.role == 'admin',
        stages=MILESTONE_STAGES,
        risk_categories=RISK_CATEGORIES,
        today=datetime.now().strftime('%Y-%m-%d'),
    )


@app.route('/risks')
@login_required
def risks_page():
    q_case = (request.args.get('case') or '').strip()
    q_cat = (request.args.get('cat') or '').strip()
    q_level = (request.args.get('level') or '').strip()
    q_status = (request.args.get('status') or '').strip()

    sql = "SELECT * FROM risk_items WHERE 1=1"
    params = []

    if current_user.role == 'employee':
        sql += " AND employee_id = ?"
        params.append(current_user.id)

    if q_case:
        sql += " AND case_id = ?"
        params.append(q_case)
    if q_cat:
        sql += " AND category = ?"
        params.append(q_cat)
    if q_level:
        sql += " AND level = ?"
        params.append(q_level)
    if q_status:
        sql += " AND status = ?"
        params.append(q_status)

    sql += " ORDER BY id DESC LIMIT 300"

    with get_db() as db:
        rows = db.execute(sql, tuple(params)).fetchall()
        case_catalog = get_case_catalog(db)

        # 統計（供頁面頂部顯示）
        stats = db.execute(
            """
            SELECT
              SUM(CASE WHEN status!='closed' THEN 1 ELSE 0 END) AS open_total,
              SUM(CASE WHEN status!='closed' AND level='high' THEN 1 ELSE 0 END) AS open_high,
              SUM(CASE WHEN status!='closed' AND level='medium' THEN 1 ELSE 0 END) AS open_medium,
              SUM(CASE WHEN status!='closed' AND level='low' THEN 1 ELSE 0 END) AS open_low
            FROM risk_items
            """
        ).fetchone()

    today = datetime.now().strftime('%Y-%m-%d')

    return render_template(
        'risks.html',
        rows=rows,
        cases=case_catalog,
        categories=RISK_CATEGORIES,
        is_admin=current_user.role == 'admin',
        q_case=q_case,
        q_cat=q_cat,
        q_level=q_level,
        q_status=q_status,
        stats=stats,
        today=today,
    )


@app.route('/tasks')
@login_required
def tasks_page():
    q_case = (request.args.get('case') or '').strip()
    q_status = (request.args.get('status') or '').strip()
    q_owner = (request.args.get('owner') or '').strip()

    sql = "SELECT * FROM tasks WHERE 1=1"
    params = []

    if current_user.role == 'employee':
        sql += " AND owner_id = ?"
        params.append(current_user.id)
    else:
        if q_owner:
            sql += " AND owner_id = ?"
            params.append(q_owner)

    if q_case:
        sql += " AND case_id = ?"
        params.append(q_case)

    if q_status:
        sql += " AND status = ?"
        params.append(q_status)

    sql += " ORDER BY CASE status WHEN 'open' THEN 0 WHEN 'in_progress' THEN 1 ELSE 2 END, due_date DESC, id DESC LIMIT 300"

    today = datetime.now().strftime('%Y-%m-%d')

    with get_db() as db:
        rows = db.execute(sql, tuple(params)).fetchall()
        case_catalog = get_case_catalog(db)

        stats = db.execute(
            """
            SELECT
              SUM(CASE WHEN status!='closed' THEN 1 ELSE 0 END) AS open_total,
              SUM(CASE WHEN status!='closed' AND due_date!='' AND due_date < ? THEN 1 ELSE 0 END) AS overdue
            FROM tasks
            """,
            (today,)
        ).fetchone()

    owners = [{'id': 'admin', 'name': '管理員'}] + list(EMPLOYEES)
    return render_template('tasks.html', rows=rows, cases=case_catalog, owners=owners,
                           is_admin=current_user.role == 'admin', q_case=q_case, q_status=q_status, q_owner=q_owner,
                           stats=stats, today=today)


@app.route('/tasks/create', methods=['POST'])
@login_required
def tasks_create():
    if current_user.role != 'admin':
        abort(403)

    case_id = (request.form.get('case_id') or '').strip()
    title = (request.form.get('title') or '').strip()
    description = (request.form.get('description') or '').strip()
    priority = (request.form.get('priority') or 'medium').strip()
    owner_id = (request.form.get('owner_id') or '').strip()
    due_date = (request.form.get('due_date') or '').strip()
    related_risk_id = (request.form.get('related_risk_id') or '').strip()

    if not case_id or not title:
        flash('請填寫案場與任務標題', 'error')
        return redirect(request.headers.get('Referer') or url_for('tasks_page'))

    if priority not in ('high', 'medium', 'low'):
        priority = 'medium'

    rr = None
    if related_risk_id:
        try:
            rr = int(related_risk_id)
        except Exception:
            rr = None

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        db.execute(
            """
            INSERT INTO tasks(case_id, title, description, priority, status, owner_id, due_date, related_risk_id,
                              created_at, created_by, updated_at, updated_by)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (case_id, title, description, priority, 'open', owner_id, due_date, rr, now, current_user.id, now, current_user.id)
        )
        db.commit()

    audit('task_create', actor_id=current_user.id, target_id=case_id, detail={'title': title, 'owner': owner_id, 'due': due_date, 'priority': priority})
    flash('✅ 任務已建立', 'success')
    return redirect(request.headers.get('Referer') or url_for('tasks_page'))


@app.route('/tasks/update/<int:task_id>', methods=['POST'])
@login_required
def tasks_update(task_id: int):
    with get_db() as db:
        row = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
        if not row:
            flash('任務不存在', 'error')
            return redirect(url_for('tasks_page'))

        # 管理員可改全部；員工只能改自己的狀態
        if current_user.role != 'admin' and row['owner_id'] != current_user.id:
            abort(403)

        status = (request.form.get('status') or row['status'] or 'open').strip()
        if status not in ('open', 'in_progress', 'closed'):
            status = row['status']

        owner_id = row['owner_id']
        due_date = row['due_date']
        priority = row['priority']

        if current_user.role == 'admin':
            owner_id = (request.form.get('owner_id') or row['owner_id'] or '').strip()
            due_date = (request.form.get('due_date') or row['due_date'] or '').strip()
            priority = (request.form.get('priority') or row['priority'] or 'medium').strip()
            if priority not in ('high', 'medium', 'low'):
                priority = row['priority']

        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db.execute(
            'UPDATE tasks SET status=?, owner_id=?, due_date=?, priority=?, updated_at=?, updated_by=? WHERE id=?',
            (status, owner_id, due_date, priority, now, current_user.id, task_id)
        )
        db.commit()

    audit('task_update', actor_id=current_user.id, target_id=str(task_id), detail={'status': status, 'owner': owner_id, 'due': due_date, 'priority': priority})
    flash('✅ 任務已更新', 'success')
    return redirect(request.headers.get('Referer') or url_for('tasks_page'))


@app.route('/risks/update/<int:risk_id>', methods=['POST'])
@login_required
def risks_update(risk_id: int):
    with get_db() as db:
        row = db.execute('SELECT * FROM risk_items WHERE id = ?', (risk_id,)).fetchone()
        if not row:
            flash('風險不存在', 'error')
            return redirect(url_for('risks_page'))

        if current_user.role != 'admin' and row['employee_id'] != current_user.id:
            abort(403)

        status = (request.form.get('status') or row['status'] or 'open').strip()
        owner_id = (request.form.get('owner_id') or row['owner_id'] or '').strip()
        due_date = (request.form.get('due_date') or row['due_date'] or '').strip()
        category = (request.form.get('category') or row['category'] or '協調').strip()
        level = (request.form.get('level') or row['level'] or 'low').strip()

        if category not in RISK_CATEGORIES:
            category = '其他'
        if level not in ('high', 'medium', 'low'):
            level = 'low'
        if status not in ('open', 'in_progress', 'closed'):
            status = 'open'

        db.execute(
            'UPDATE risk_items SET status=?, owner_id=?, due_date=?, category=?, level=? WHERE id=?',
            (status, owner_id, due_date, category, level, risk_id)
        )
        db.commit()

    audit('risk_update', actor_id=current_user.id, target_id=str(risk_id), detail={'status': status, 'owner': owner_id, 'due': due_date, 'cat': category, 'level': level})
    flash('✅ 風險已更新', 'success')

    # 回到上一頁
    ref = request.headers.get('Referer')
    return redirect(ref or url_for('risks_page'))


@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})


# ===================== API =====================
@app.route('/api/status')
@login_required
def api_status():
    today = datetime.now().strftime('%Y-%m-%d')
    today_dir = REPORTS_DIR / today
    submitted = len(list(today_dir.glob('*_report.json'))) if today_dir.exists() else 0
    return jsonify({
        'status': 'ok',
        'date': today,
        'total_employees': len(EMPLOYEES),
        'submitted': submitted,
        'rate': round(submitted / len(EMPLOYEES) * 100, 1)
    })

@app.route('/api/cases')
@login_required
def api_cases():
    return jsonify(get_case_catalog())


# ===================== AI 讀取 API（管理員） =====================
@app.route('/api/ai/overview')
@login_required
def api_ai_overview():
    """提供 AI 使用的案件總覽（可用於生成風險/行動建議）。管理員限定。"""
    if current_user.role != 'admin':
        abort(403)

    today = datetime.now().strftime('%Y-%m-%d')

    with get_db() as db:
        out_cases = []
        case_catalog = get_case_catalog(db)
        for c in case_catalog:
            cid = c['id']
            st = db.execute('SELECT * FROM case_status WHERE case_id=?', (cid,)).fetchone()

            risk = db.execute(
                """
                SELECT
                  SUM(CASE WHEN status!='closed' THEN 1 ELSE 0 END) AS open_total,
                  SUM(CASE WHEN status!='closed' AND level='high' THEN 1 ELSE 0 END) AS open_high,
                  SUM(CASE WHEN status!='closed' AND level='medium' THEN 1 ELSE 0 END) AS open_medium,
                  SUM(CASE WHEN status!='closed' AND level='low' THEN 1 ELSE 0 END) AS open_low
                FROM risk_items WHERE case_id=?
                """,
                (cid,)
            ).fetchone()

            task = db.execute(
                """
                SELECT
                  SUM(CASE WHEN status!='closed' THEN 1 ELSE 0 END) AS open_total,
                  SUM(CASE WHEN status='open' THEN 1 ELSE 0 END) AS open_count,
                  SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END) AS in_progress_count
                FROM tasks WHERE case_id=?
                """,
                (cid,)
            ).fetchone()

            last_work = db.execute(
                """
                SELECT MAX(r.report_date) AS last_date
                FROM work_items w JOIN reports r ON w.report_key=r.report_key
                WHERE w.case_id=?
                """,
                (cid,)
            ).fetchone()

            out_cases.append({
                'case_id': cid,
                'case_name': c.get('name', cid),
                'case_type': c.get('type', ''),
                'milestone_stage': (st['stage'] if st else ''),
                'progress_percent': (st['percent'] if st and st['percent'] is not None else None),
                'status_note': (st['note'] if st else ''),
                'status_updated_at': (st['updated_at'] if st else ''),
                'last_work_date': last_work['last_date'] if last_work else '',
                'open_risks': {
                    'total': risk['open_total'] or 0,
                    'high': risk['open_high'] or 0,
                    'medium': risk['open_medium'] or 0,
                    'low': risk['open_low'] or 0,
                },
                'open_tasks': {
                    'total': task['open_total'] or 0,
                    'open': task['open_count'] or 0,
                    'in_progress': task['in_progress_count'] or 0,
                }
            })

        overdue = db.execute(
            """
            SELECT * FROM tasks
            WHERE status!='closed' AND due_date!='' AND due_date < ?
            ORDER BY due_date ASC, id ASC
            LIMIT 200
            """,
            (today,)
        ).fetchall()

        overdue_tasks = [dict(r) for r in overdue]

    return jsonify({
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'cases': out_cases,
        'overdue_tasks': overdue_tasks,
    })


@app.route('/api/ai/case/<case_id>')
@login_required
def api_ai_case(case_id):
    """提供單一案場的工作/風險/任務明細（管理員限定）。"""
    if current_user.role != 'admin':
        abort(403)

    with get_db() as db:
        st = db.execute('SELECT * FROM case_status WHERE case_id=?', (case_id,)).fetchone()
        works = db.execute(
            """
            SELECT r.report_date, r.employee_id, w.*
            FROM work_items w JOIN reports r ON w.report_key=r.report_key
            WHERE w.case_id=?
            ORDER BY r.report_date DESC, w.id DESC
            LIMIT 80
            """,
            (case_id,)
        ).fetchall()

        plans = db.execute(
            """
            SELECT r.report_date, r.employee_id, p.*
            FROM plan_items p JOIN reports r ON p.report_key=r.report_key
            WHERE p.case_id=?
            ORDER BY r.report_date DESC, p.id DESC
            LIMIT 80
            """,
            (case_id,)
        ).fetchall()
        risks = db.execute(
            "SELECT * FROM risk_items WHERE case_id=? ORDER BY id DESC LIMIT 120",
            (case_id,)
        ).fetchall()
        tasks = db.execute(
            "SELECT * FROM tasks WHERE case_id=? ORDER BY CASE status WHEN 'open' THEN 0 WHEN 'in_progress' THEN 1 ELSE 2 END, due_date DESC, id DESC LIMIT 200",
            (case_id,)
        ).fetchall()

        # 附件：以「有關聯到此案場的 report_key」回推出該些日報的附件
        keys = set([r['report_key'] for r in db.execute('SELECT DISTINCT report_key FROM work_items WHERE case_id=?', (case_id,)).fetchall()])
        keys |= set([r['report_key'] for r in db.execute('SELECT DISTINCT report_key FROM plan_items WHERE case_id=?', (case_id,)).fetchall()])
        keys |= set([r['report_key'] for r in db.execute('SELECT DISTINCT report_key FROM risk_items WHERE case_id=?', (case_id,)).fetchall()])

        attachments = []
        if keys:
            ph = ','.join(['?'] * len(keys))
            attachments = db.execute(
                f'SELECT * FROM attachments WHERE report_key IN ({ph}) ORDER BY id DESC LIMIT 200',
                tuple(keys)
            ).fetchall()

    return jsonify({
        'case_id': case_id,
        'status': dict(st) if st else {},
        'works': [dict(r) for r in works],
        'plans': [dict(r) for r in plans],
        'risks': [dict(r) for r in risks],
        'tasks': [dict(r) for r in tasks],
        'attachments': [dict(r) for r in attachments],
    })

# ===================== Excel 匯出 =====================
@app.route('/export/excel')
@login_required
def export_excel():
    if current_user.role != 'admin':
        flash('無權限', 'error'); return redirect(url_for('index'))
    if not EXCEL_OK:
        flash('伺服器未安裝 openpyxl，無法匯出', 'error')
        return redirect(url_for('admin_dashboard'))

    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    date_dir = REPORTS_DIR / date_param
    reports = []
    if date_dir.exists():
        for f in sorted(date_dir.glob('*_report.json')):
            reports.append(json.loads(f.read_text(encoding='utf-8')))

    wb = openpyxl.Workbook()

    # ── 工作表一：摘要 ──
    ws_sum = wb.active
    ws_sum.title = '日報摘要'
    header_fill = PatternFill(fill_type='solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    def hcell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
        if width: ws.column_dimensions[c.column_letter].width = width
        return c
    def dcell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = thin; return c

    ws_sum.row_dimensions[1].height = 30
    hcell(ws_sum,1,1,'員工姓名',15); hcell(ws_sum,1,2,'部門',12)
    hcell(ws_sum,1,3,'提交時間',18); hcell(ws_sum,1,4,'工作項數',10)
    hcell(ws_sum,1,5,'計畫項數',10); hcell(ws_sum,1,6,'風險項數',10)
    hcell(ws_sum,1,7,'高風險數',10); hcell(ws_sum,1,8,'今日收穫',30)

    # 未提交員工
    submitted_ids = {r['employee_id'] for r in reports}
    for emp in EMPLOYEES:
        if emp['id'] not in submitted_ids:
            row = ws_sum.max_row + 1
            dcell(ws_sum, row, 1, emp['name'])
            dcell(ws_sum, row, 2, '未提交')
            for col in range(3, 9): dcell(ws_sum, row, col, '-')
            ws_sum.cell(row=row, column=2).font = Font(color='FF0000')

    for r in reports:
        high_count = sum(1 for ri in r.get('risk_items',[]) if ri.get('risk_level')=='high')
        row = ws_sum.max_row + 1
        dcell(ws_sum, row, 1, r['employee_name'])
        dcell(ws_sum, row, 2, r.get('department',''))
        dcell(ws_sum, row, 3, r['submit_time'])
        dcell(ws_sum, row, 4, len(r.get('work_items',[])))
        dcell(ws_sum, row, 5, len(r.get('plan_items',[])))
        dcell(ws_sum, row, 6, len(r.get('risk_items',[])))
        dcell(ws_sum, row, 7, high_count)
        dcell(ws_sum, row, 8, r.get('summary',{}).get('today_gain',''))
        if high_count > 0:
            ws_sum.cell(row=row, column=7).fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
            ws_sum.cell(row=row, column=7).font = Font(color='991B1B', bold=True)

    # ── 工作表二：工作明細 ──
    ws_work = wb.create_sheet('工作明細')
    headers = ['員工','案場','工作類型','工作內容','進度%','工時','產出物','狀態']
    for i, h in enumerate(headers, 1):
        hcell(ws_work, 1, i, h, [12,15,10,40,8,8,20,10][i-1])
    ws_work.row_dimensions[1].height = 25
    for r in reports:
        for w in r.get('work_items', []):
            row = ws_work.max_row + 1
            dcell(ws_work, row, 1, r['employee_name'])
            dcell(ws_work, row, 2, w.get('case_id',''))
            dcell(ws_work, row, 3, w.get('work_type',''))
            dcell(ws_work, row, 4, w.get('work_content',''))
            dcell(ws_work, row, 5, w.get('progress',''))
            dcell(ws_work, row, 6, w.get('hours',''))
            dcell(ws_work, row, 7, w.get('output',''))
            dcell(ws_work, row, 8, w.get('status',''))

    # ── 工作表三：風險清單 ──
    ws_risk = wb.create_sheet('風險清單')
    rh = ['員工','案場','風險等級','風險說明','影響層面','需要協助']
    for i, h in enumerate(rh, 1):
        hcell(ws_risk, 1, i, h, [12,15,10,35,25,25][i-1])
    ws_risk.row_dimensions[1].height = 25
    risk_colors = {'high':'FEE2E2', 'medium':'FEF3C7', 'low':'DCFCE7'}
    for r in reports:
        for ri in r.get('risk_items', []):
            row = ws_risk.max_row + 1
            dcell(ws_risk, row, 1, r['employee_name'])
            dcell(ws_risk, row, 2, ri.get('case_id',''))
            lv = ri.get('risk_level','low')
            lv_text = {'high':'🔴 高風險','medium':'🟠 中風險','low':'🟢 低風險'}.get(lv, lv)
            c = dcell(ws_risk, row, 3, lv_text)
            c.fill = PatternFill(fill_type='solid', fgColor=risk_colors.get(lv,'FFFFFF'))
            dcell(ws_risk, row, 4, ri.get('risk_desc',''))
            dcell(ws_risk, row, 5, ri.get('risk_impact',''))
            dcell(ws_risk, row, 6, ri.get('need_help',''))

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    audit('export_excel', actor_id=current_user.id, detail={'date': date_param})
    filename = f"日報_{date_param}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ===================== 員工管理頁 =====================
@app.route('/admin/reset-password/<user_id>', methods=['POST'])
@login_required
def admin_reset_password(user_id):
    if current_user.role != 'admin':
        flash('無權限', 'error')
        return redirect(url_for('index'))

    if user_id not in USERS:
        flash('員工不存在', 'error')
        return redirect(url_for('admin_employees'))

    # 產生一次性臨時密碼（不在登入頁顯示；只回給管理員）
    import secrets
    temp_pw = secrets.token_urlsafe(9)  # 約 12~13 字元

    u = USERS[user_id]
    u['password_hash'] = generate_password_hash(temp_pw)
    u['password_temp'] = True
    USERS[user_id] = u
    save_users(USERS)
    refresh_employee_cache()

    audit('admin_reset_password', actor_id=current_user.id, target_id=user_id)
    flash(f'已重設 {u.get("name")} 的密碼。臨時密碼：{temp_pw}（請轉交本人並提醒登入後立即修改）', 'success')
    return redirect(url_for('admin_employees'))


@app.route('/admin/employees')
@login_required
def admin_employees():
    if current_user.role != 'admin':
        flash('無權限', 'error'); return redirect(url_for('index'))

    # 近 30 天提交統計
    today = datetime.now()
    stats = {}
    for emp in EMPLOYEES:
        stats[emp['id']] = {'name': emp['name'], 'total_days': 0, 'submit_days': 0, 'avg_work': 0}

    total_work_by_emp = {e['id']: [] for e in EMPLOYEES}
    for i in range(30):
        d = (today - timedelta(days=i)).strftime('%Y-%m-%d')
        dd = REPORTS_DIR / d
        if not dd.exists(): continue
        for emp in EMPLOYEES:
            f = dd / f"{emp['id']}_report.json"
            stats[emp['id']]['total_days'] += 1
            if f.exists():
                stats[emp['id']]['submit_days'] += 1
                r = json.loads(f.read_text(encoding='utf-8'))
                total_work_by_emp[emp['id']].append(len(r.get('work_items', [])))

    for eid, wlist in total_work_by_emp.items():
        stats[eid]['avg_work'] = round(sum(wlist)/len(wlist), 1) if wlist else 0
        sd = stats[eid]['submit_days']
        td = stats[eid]['total_days']
        stats[eid]['rate'] = round(sd/td*100, 1) if td else 0

    return render_template('admin_employees.html',
        employees=EMPLOYEES, stats=stats)

# ===================== 逾時未交提醒 =====================
@app.route('/api/reminder/check')
@login_required
def reminder_check():
    """檢查今天哪些員工還未提交，回傳名單（admin only）"""
    if current_user.role != 'admin':
        return jsonify({'error': '無權限'}), 403
    today = datetime.now().strftime('%Y-%m-%d')
    today_dir = REPORTS_DIR / today
    submitted_ids = set()
    if today_dir.exists():
        for f in today_dir.glob('*_report.json'):
            submitted_ids.add(f.stem.replace('_report', ''))
    missing = [{'id': e['id'], 'name': e['name']} for e in EMPLOYEES if e['id'] not in submitted_ids]
    hour = datetime.now().hour
    overdue = hour >= 18  # 18:00 後算逾時
    return jsonify({
        'date': today,
        'total': len(EMPLOYEES),
        'submitted': len(submitted_ids),
        'missing': missing,
        'overdue': overdue,
        'current_hour': hour
    })

# ===================== 編輯留痕 =====================
def append_edit_log(report_file: Path, operator: str, action: str):
    """在日報 JSON 中追加編輯記錄"""
    if not report_file.exists():
        return
    with open(report_file, encoding='utf-8') as f:
        data = json.load(f)
    if 'edit_log' not in data:
        data['edit_log'] = []
    data['edit_log'].append({
        'operator': operator,
        'action': action,
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ===================== 圖表資料 API =====================
@app.route('/api/chart/weekly')
@login_required
def chart_weekly():
    """最近 7 天各日提交率（供圖表使用）"""
    if current_user.role != 'admin':
        return jsonify({'error': '無權限'}), 403
    from datetime import timedelta
    result = []
    total = len(EMPLOYEES)
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        ddir = REPORTS_DIR / d
        count = len(list(ddir.glob('*_report.json'))) if ddir.exists() else 0
        result.append({'date': d, 'submitted': count, 'total': total,
                        'rate': round(count / total * 100, 1) if total else 0})
    return jsonify(result)

@app.route('/api/chart/risk_summary')
@login_required
def chart_risk_summary():
    """今日風險分佈統計"""
    if current_user.role != 'admin':
        return jsonify({'error': '無權限'}), 403
    today = datetime.now().strftime('%Y-%m-%d')
    today_dir = REPORTS_DIR / today
    high = medium = low = 0
    if today_dir.exists():
        for f in today_dir.glob('*_report.json'):
            with open(f, encoding='utf-8') as fp:
                r = json.load(fp)
            for risk in r.get('risk_items', []):
                lv = risk.get('risk_level', 'low')
                if lv == 'high': high += 1
                elif lv == 'medium': medium += 1
                else: low += 1
    return jsonify({'high': high, 'medium': medium, 'low': low})

# ===================== 啟動 =====================
if __name__ == '__main__':
    print("=" * 60)
    print("🏢 昱金生能源集團 - 員工日報系統 v3.0")
    print("=" * 60)
    print(f"🌐 http://127.0.0.1:5000")
    print("🔐 管理員帳號：admin（密碼請由管理員另行提供/保存）")
    print("👷 員工初始密碼：首次登入需修改（密碼不在此顯示）")
    print("=" * 60)
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)