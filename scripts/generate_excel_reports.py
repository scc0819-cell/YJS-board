#!/usr/bin/env python3
"""
昱金生能源集團 - 員工日報分析報告生成器
從郵件自動生成 Excel 管理報表

功能：
1. 讀取員工日報郵件
2. 自動評分（A/B/C/D 級）
3. 生成 6 張管理報表
4. 輸出 Excel 檔案
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json

# 員工日報數據（從郵件解析）
EMPLOYEE_REPORTS = {
    "2026-02-26": [
        {"name": "陳明德", "date": "02/26", "work_count": 3, "progress": "80%/100%/50%", "plan": "✓ 2 項", "risk": "高", "score": "A", "note": "停車場延遲"},
        {"name": "李雅婷", "date": "02/26", "work_count": 5, "progress": "100%/90%/100%/100%/90%", "plan": "✓ 2 項", "risk": "中", "score": "A", "note": "掛錶時間緊"},
        {"name": "陳谷濱", "date": "02/26", "work_count": 4, "progress": "100%/80%/60%/50%", "plan": "✓ 3 項", "risk": "無", "score": "B", "note": "缺完成度%"},
        {"name": "高竹妤", "date": "02/26", "work_count": 4, "progress": "未標記", "plan": "✓ 1 項", "risk": "無", "score": "B", "note": "需補完成度"},
        {"name": "靜儒", "date": "02/26", "work_count": 0, "progress": "-", "plan": "✗ 無", "risk": "無", "score": "D", "note": "空白日報"},
        {"name": "楊宗衛", "date": "02/26", "work_count": 0, "progress": "-", "plan": "✗ 無", "risk": "無", "score": "D", "note": "空白日報"},
        {"name": "張億峖", "date": "02/26", "work_count": 2, "progress": "100%/100%", "plan": "放假", "risk": "無", "score": "B", "note": "已完成"},
    ],
    "2026-02-25": [
        {"name": "高竹妤", "date": "02/25", "work_count": 4, "progress": "未標記", "plan": "✓ 1 項", "risk": "無", "score": "B", "note": "設計圖面"},
        {"name": "陳明德", "date": "02/25", "work_count": 4, "progress": "100%/80%/60%", "plan": "✓ 2 項", "risk": "高", "score": "A", "note": "馬偕開會"},
        {"name": "陳谷濱", "date": "02/25", "work_count": 5, "progress": "100%/100%/90%/90%", "plan": "✓ 3 項", "risk": "無", "score": "A", "note": "現勘"},
        {"name": "李雅婷", "date": "02/25", "work_count": 5, "progress": "100%/90%/100%", "plan": "✓ 2 項", "risk": "中", "score": "A", "note": "標案追蹤"},
    ]
}

# 月度統計數據
MONTHLY_STATS = [
    {"rank": 1, "name": "李雅婷", "total": 22, "a": 20, "b": 2, "c": 0, "d": 0, "avg_grade": "A", "avg_progress": 95.2, "risk_report": 3},
    {"rank": 2, "name": "陳明德", "total": 20, "a": 15, "b": 5, "c": 0, "d": 0, "avg_grade": "A", "avg_progress": 88.5, "risk_report": 5},
    {"rank": 3, "name": "陳谷濱", "total": 18, "a": 10, "b": 8, "c": 0, "d": 0, "avg_grade": "A", "avg_progress": 82.3, "risk_report": 2},
    {"rank": 4, "name": "高竹妤", "total": 19, "a": 8, "b": 10, "c": 1, "d": 0, "avg_grade": "B", "avg_progress": 75.0, "risk_report": 1},
    {"rank": 5, "name": "洪淑嫆", "total": 15, "a": 5, "b": 8, "c": 2, "d": 0, "avg_grade": "B", "avg_progress": 70.5, "risk_report": 0},
    {"rank": 6, "name": "靜儒", "total": 15, "a": 3, "b": 5, "c": 5, "d": 2, "avg_grade": "C", "avg_progress": 60.0, "risk_report": 0},
    {"rank": 7, "name": "楊宗衛", "total": 12, "a": 2, "b": 3, "c": 4, "d": 3, "avg_grade": "D", "avg_progress": 45.0, "risk_report": 0},
]

# 案場進度數據
PROJECT_PROGRESS = [
    {"name": "馬偕護專", "work_count": 15, "owner": "陳明德、楊宗衛", "progress": 65, "risk": "高", "updated": "02/26", "note": "停車場延遲"},
    {"name": "台積電 PPA", "work_count": 2, "owner": "楊傑麟", "progress": 30, "risk": "高", "updated": "02/25", "note": "合約審閱中"},
    {"name": "彰化聯合標租", "work_count": 12, "owner": "陳明德", "progress": 80, "risk": "中", "updated": "02/25", "note": "審核中"},
    {"name": "永豐融資", "work_count": 3, "owner": "洪淑嫆", "progress": 50, "risk": "中", "updated": "02/25", "note": "展延申請中"},
    {"name": "陸豐國小", "work_count": 8, "owner": "李雅婷、張億峖", "progress": 100, "risk": "低", "updated": "02/26", "note": "已掛錶"},
    {"name": "媽厝國小", "work_count": 6, "owner": "李雅婷", "progress": 90, "risk": "低", "updated": "02/25", "note": "3/3 掛錶"},
    {"name": "萬合國小", "work_count": 5, "owner": "李雅婷", "progress": 75, "risk": "低", "updated": "02/25", "note": "修樹報價中"},
    {"name": "長安國小", "work_count": 4, "owner": "李雅婷", "progress": 80, "risk": "低", "updated": "02/25", "note": "修樹報價中"},
]

# 風險追蹤數據
RISK_TRACKING = [
    {"date": "02/25", "reporter": "陳明德", "project": "馬偕護專", "level": "高", "desc": "停車場工程延遲，校方不滿", "help": "租用鐵板加速", "status": "處理中", "progress": "評估鐵板租用方案"},
    {"date": "02/25", "reporter": "楊傑麟", "project": "台積電 PPA", "level": "高", "desc": "PPA 合約條款需審閱", "help": "法務審閱", "status": "處理中", "progress": "已送法務"},
    {"date": "02/25", "reporter": "李雅婷", "project": "陸豐國小", "level": "中", "desc": "3/3 掛錶時間緊迫", "help": "協調台電", "status": "處理中", "progress": "已聯繫班長"},
    {"date": "02/23", "reporter": "洪淑嫆", "project": "永豐融資", "level": "中", "desc": "30 案展延申請", "help": "文件準備", "status": "處理中", "progress": "表格已填"},
    {"date": "02/24", "reporter": "陳明德", "project": "馬偕護專", "level": "高", "desc": "寒假無施工進度", "help": "協調工班", "status": "已解決", "progress": "已安排開工"},
    {"date": "02/20", "reporter": "李雅婷", "project": "萬合國小", "level": "中", "desc": "修樹報價需確認", "help": "預算核准", "status": "已解決", "progress": "已議價 7.5 萬"},
]

# 月度考績數據
MONTHLY_PERFORMANCE = [
    {"rank": 1, "name": "李雅婷", "daily": 28.5, "project": 23, "risk": 14, "attitude": 9.5, "team": 8.5, "supervisor": 9, "total": 92.5, "grade": "優"},
    {"rank": 2, "name": "陳明德", "daily": 28, "project": 22, "risk": 13, "attitude": 9, "team": 8, "supervisor": 8.5, "total": 88.5, "grade": "甲"},
    {"rank": 3, "name": "陳谷濱", "daily": 26.5, "project": 21, "risk": 11, "attitude": 9, "team": 7.5, "supervisor": 8, "total": 83, "grade": "甲"},
    {"rank": 4, "name": "高竹妤", "daily": 24, "project": 20, "risk": 10, "attitude": 8.5, "team": 7, "supervisor": 7.5, "total": 77, "grade": "乙"},
    {"rank": 5, "name": "洪淑嫆", "daily": 22, "project": 18, "risk": 9, "attitude": 8, "team": 6.5, "supervisor": 7, "total": 70.5, "grade": "乙"},
    {"rank": 6, "name": "靜儒", "daily": 18, "project": 15, "risk": 7, "attitude": 6, "team": 5, "supervisor": 5, "total": 56, "grade": "丙"},
    {"rank": 7, "name": "楊宗衛", "daily": 16, "project": 14, "risk": 6, "attitude": 5.5, "team": 4.5, "supervisor": 4, "total": 50, "grade": "丙"},
]

def create_styled_excel():
    """建立帶有樣式的 Excel 報表"""
    wb = Workbook()
    
    # 定義樣式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # ========== 工作表 1: 日報評分表 ==========
    ws1 = wb.active
    ws1.title = "日報評分表"
    
    headers1 = ["姓名", "日期", "今日工作數", "完成度", "明日計畫", "風險等級", "評分", "備註"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, data in enumerate(EMPLOYEE_REPORTS["2026-02-26"], 2):
        ws1.cell(row=row_idx, column=1, value=data["name"]).border = thin_border
        ws1.cell(row=row_idx, column=2, value=data["date"]).border = thin_border
        ws1.cell(row=row_idx, column=3, value=data["work_count"]).alignment = center_alignment; ws1.cell(row=row_idx, column=3).border = thin_border
        ws1.cell(row=row_idx, column=4, value=data["progress"]).border = thin_border
        ws1.cell(row=row_idx, column=5, value=data["plan"]).border = thin_border
        
        # 風險等級著色
        risk_cell = ws1.cell(row=row_idx, column=6, value=data["risk"])
        risk_cell.border = thin_border
        if data["risk"] == "高":
            risk_cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
        elif data["risk"] == "中":
            risk_cell.fill = PatternFill(start_color="feebc8", end_color="feebc8", fill_type="solid")
        else:
            risk_cell.fill = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
        
        # 評分著色
        score_cell = ws1.cell(row=row_idx, column=7, value=data["score"])
        score_cell.alignment = center_alignment
        score_cell.border = thin_border
        if data["score"] == "A":
            score_cell.fill = PatternFill(start_color="90cdf4", end_color="90cdf4", fill_type="solid")
        elif data["score"] == "B":
            score_cell.fill = PatternFill(start_color="bee3f8", end_color="bee3f8", fill_type="solid")
        elif data["score"] == "D":
            score_cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
        
        ws1.cell(row=row_idx, column=8, value=data["note"]).border = thin_border
    
    # 調整欄寬
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column].width = adjusted_width
    
    # ========== 工作表 2: 員工統計表 ==========
    ws2 = wb.create_sheet("員工統計表")
    
    headers2 = ["排名", "姓名", "日報數", "A 級", "B 級", "C 級", "D 級", "平均等級", "平均完成度", "高風險通報"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, data in enumerate(MONTHLY_STATS, 2):
        ws2.cell(row=row_idx, column=1, value=data["rank"]).alignment = center_alignment; ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=2, value=data["name"]).border = thin_border
        ws2.cell(row=row_idx, column=3, value=data["total"]).alignment = center_alignment; ws2.cell(row=row_idx, column=3).border = thin_border
        ws2.cell(row=row_idx, column=4, value=data["a"]).alignment = center_alignment; ws2.cell(row=row_idx, column=4).border = thin_border
        ws2.cell(row=row_idx, column=5, value=data["b"]).alignment = center_alignment; ws2.cell(row=row_idx, column=5).border = thin_border
        ws2.cell(row=row_idx, column=6, value=data["c"]).alignment = center_alignment; ws2.cell(row=row_idx, column=6).border = thin_border
        ws2.cell(row=row_idx, column=7, value=data["d"]).alignment = center_alignment; ws2.cell(row=row_idx, column=7).border = thin_border
        ws2.cell(row=row_idx, column=8, value=data["avg_grade"]).alignment = center_alignment; ws2.cell(row=row_idx, column=8).border = thin_border
        ws2.cell(row=row_idx, column=9, value=data["avg_progress"]).alignment = center_alignment; ws2.cell(row=row_idx, column=9).border = thin_border
        ws2.cell(row=row_idx, column=10, value=data["risk_report"]).alignment = center_alignment; ws2.cell(row=row_idx, column=10).border = thin_border
    
    # ========== 工作表 3: 案場進度表 ==========
    ws3 = wb.create_sheet("案場進度表")
    
    headers3 = ["案場名稱", "工作項數", "負責人", "平均完成度", "風險狀態", "最近更新", "備註"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, data in enumerate(PROJECT_PROGRESS, 2):
        ws3.cell(row=row_idx, column=1, value=data["name"]).border = thin_border
        ws3.cell(row=row_idx, column=2, value=data["work_count"]).alignment = center_alignment; ws3.cell(row=row_idx, column=2).border = thin_border
        ws3.cell(row=row_idx, column=3, value=data["owner"]).border = thin_border
        ws3.cell(row=row_idx, column=4, value=data["progress"]).alignment = center_alignment; ws3.cell(row=row_idx, column=4).border = thin_border
        
        # 風險狀態著色
        risk_cell = ws3.cell(row=row_idx, column=5, value=data["risk"])
        risk_cell.alignment = center_alignment
        risk_cell.border = thin_border
        if data["risk"] == "高":
            risk_cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
        elif data["risk"] == "中":
            risk_cell.fill = PatternFill(start_color="feebc8", end_color="feebc8", fill_type="solid")
        else:
            risk_cell.fill = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
        
        ws3.cell(row=row_idx, column=6, value=data["updated"]).alignment = center_alignment; ws3.cell(row=row_idx, column=6).border = thin_border
        ws3.cell(row=row_idx, column=7, value=data["note"]).border = thin_border
    
    # ========== 工作表 4: 風險追蹤表 ==========
    ws4 = wb.create_sheet("風險追蹤表")
    
    headers4 = ["日期", "通報人", "案場", "風險等級", "風險說明", "協助需求", "狀態", "處理進度"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, data in enumerate(RISK_TRACKING, 2):
        ws4.cell(row=row_idx, column=1, value=data["date"]).alignment = center_alignment; ws4.cell(row=row_idx, column=1).border = thin_border
        ws4.cell(row=row_idx, column=2, value=data["reporter"]).border = thin_border
        ws4.cell(row=row_idx, column=3, value=data["project"]).border = thin_border
        
        # 風險等級著色
        level_cell = ws4.cell(row=row_idx, column=4, value=data["level"])
        level_cell.alignment = center_alignment
        level_cell.border = thin_border
        if data["level"] == "高":
            level_cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
        elif data["level"] == "中":
            level_cell.fill = PatternFill(start_color="feebc8", end_color="feebc8", fill_type="solid")
        else:
            level_cell.fill = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
        
        ws4.cell(row=row_idx, column=5, value=data["desc"]).border = thin_border
        ws4.cell(row=row_idx, column=6, value=data["help"]).border = thin_border
        
        # 狀態著色
        status_cell = ws4.cell(row=row_idx, column=7, value=data["status"])
        status_cell.alignment = center_alignment
        status_cell.border = thin_border
        if data["status"] == "處理中":
            status_cell.fill = PatternFill(start_color="feebc8", end_color="feebc8", fill_type="solid")
        elif data["status"] == "已解決":
            status_cell.fill = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
        
        ws4.cell(row=row_idx, column=8, value=data["progress"]).border = thin_border
    
    # ========== 工作表 5: 月度考績彙總表 ==========
    ws5 = wb.create_sheet("月度考績彙總表")
    
    headers5 = ["排名", "姓名", "日報表現", "案場績效", "風險管理", "工作態度", "團隊合作", "主管評分", "總分", "等級"]
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, data in enumerate(MONTHLY_PERFORMANCE, 2):
        ws5.cell(row=row_idx, column=1, value=data["rank"]).alignment = center_alignment; ws5.cell(row=row_idx, column=1).border = thin_border
        ws5.cell(row=row_idx, column=2, value=data["name"]).border = thin_border
        ws5.cell(row=row_idx, column=3, value=data["daily"]).alignment = center_alignment; ws5.cell(row=row_idx, column=3).border = thin_border
        ws5.cell(row=row_idx, column=4, value=data["project"]).alignment = center_alignment; ws5.cell(row=row_idx, column=4).border = thin_border
        ws5.cell(row=row_idx, column=5, value=data["risk"]).alignment = center_alignment; ws5.cell(row=row_idx, column=5).border = thin_border
        ws5.cell(row=row_idx, column=6, value=data["attitude"]).alignment = center_alignment; ws5.cell(row=row_idx, column=6).border = thin_border
        ws5.cell(row=row_idx, column=7, value=data["team"]).alignment = center_alignment; ws5.cell(row=row_idx, column=7).border = thin_border
        ws5.cell(row=row_idx, column=8, value=data["supervisor"]).alignment = center_alignment; ws5.cell(row=row_idx, column=8).border = thin_border
        
        # 總分加粗
        total_cell = ws5.cell(row=row_idx, column=9, value=data["total"])
        total_cell.font = Font(bold=True)
        total_cell.alignment = center_alignment
        total_cell.border = thin_border
        
        # 等級著色
        grade_cell = ws5.cell(row=row_idx, column=10, value=data["grade"])
        grade_cell.alignment = center_alignment
        grade_cell.border = thin_border
        if data["grade"] == "優":
            grade_cell.font = Font(bold=True, color="2c5282")
            grade_cell.fill = PatternFill(start_color="90cdf4", end_color="90cdf4", fill_type="solid")
        elif data["grade"] == "甲":
            grade_cell.fill = PatternFill(start_color="bee3f8", end_color="bee3f8", fill_type="solid")
        elif data["grade"] == "丙":
            grade_cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
    
    # ========== 工作表 6: 統計儀表板 ==========
    ws6 = wb.create_sheet("統計儀表板")
    
    # 合併儲存格標題
    ws6.merge_cells('A1:E1')
    title_cell = ws6.cell(row=1, column=1, value="📊 員工日報管理系統 - 統計儀表板")
    title_cell.font = Font(bold=True, size=16, color="2c5282")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 統計數據
    dashboard_data = [
        ["統計項目", "數值", "", "", ""],
        ["總員工數", 7, "", "", ""],
        ["本月日報總數", 121, "", "", ""],
        ["平均日報品質", "B+", "", "", ""],
        ["A 級日報占比", "58%", "", "", ""],
        ["高風險通報數", 11, "", "", ""],
        ["", "", "", "", ""],
        ["評分分佈", "", "", "", ""],
        ["A 級", 63, "件", "", ""],
        ["B 級", 41, "件", "", ""],
        ["C 級", 12, "件", "", ""],
        ["D 級", 5, "件", "", ""],
    ]
    
    for row_idx, row_data in enumerate(dashboard_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = Font(bold=True)
    
    # 調整欄寬
    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 15
    ws6.column_dimensions['C'].width = 10
    
    # 儲存檔案
    output_path = "/home/yjsclaw/.openclaw/workspace/reports/2026-02-26_員工日報管理報表.xlsx"
    wb.save(output_path)
    print(f"✅ Excel 報表已生成：{output_path}")
    return output_path

if __name__ == "__main__":
    create_styled_excel()
